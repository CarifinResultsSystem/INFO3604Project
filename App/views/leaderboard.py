from flask import Blueprint, render_template, jsonify, request
from flask_jwt_extended import jwt_required, current_user as jwt_current_user

from App.controllers import get_all_users_json
from App.database import db
from App.models import ScoreDocument, Season, Institution

import io
import os
import re
import pandas as pd
import numpy as np

leaderboard_views = Blueprint('leaderboard_views', __name__, template_folder='../templates')


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _all_inst_empty(row, inst_cols):
    for col in inst_cols:
        val = row[col]
        if pd.notna(val) and str(val).strip() != '':
            return False
    return True


def _load_df_data_only(file_data, original_filename):
    """
    Load a spreadsheet as a DataFrame using openpyxl data_only=True.

    This is critical: the spreadsheet uses Excel formulas for subtotals and
    grand totals. Without data_only=True, pandas reads those cells as raw
    formula strings like '=B5+B6+...' instead of the computed numeric value,
    which corrupts both parsing and the totals comparison.

    Returns a DataFrame whose first row (index 0) corresponds to spreadsheet
    row 2 (0-based row 1 is used as column headers).
    """
    ext = os.path.splitext(original_filename)[1].lower()
    buf = io.BytesIO(file_data)

    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook
        wb = load_workbook(buf, data_only=True, read_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        if len(rows_raw) < 2:
            return None, []
        # Row index 1 is the column-header row ("Event / Institution", inst names...)
        col_names = []
        for i, c in enumerate(rows_raw[1]):
            if c is not None:
                s = str(c).strip()
                if s.isdigit():
                    inst = db.session.get(Institution, int(s))
                    col_names.append(inst.insName if inst else s)
                elif s.startswith('Unnamed'):
                    col_names.append(f'Institution {s.split("_")[1]}')
                else:
                    col_names.append(s)
            else:
                col_names.append(f'Institution {i}')
        df = pd.DataFrame(rows_raw[2:], columns=col_names)
        return df, rows_raw
    else:
        df = pd.read_csv(buf, header=1)
        return df, None


def _parse_document(doc):
    """
    Parse a ScoreDocument and return:
      institutions, challenges, totals, calculated_totals, calculated_rankings
    """
    if not doc.fileData:
        return None

    df, rows_raw = _load_df_data_only(doc.fileData, doc.originalFilename)
    if df is None:
        return None

    # Rename first column to 'Rule' and drop repeated header rows
    df = df.rename(columns={df.columns[0]: 'Rule'})
    df = df[df['Rule'].astype(str).str.strip() != 'Event / Institution'].reset_index(drop=True)

    if df.empty or 'Rule' not in df.columns:
        return None

    institutions = [c for c in df.columns if c != 'Rule' and str(c).strip() != '']

    challenges = []
    totals = {}
    current_challenge = None
    current_event = None
    # True immediately after an all-caps challenge section header OR after an
    # 'Event Win Points' row — signals that the NEXT data row is an event
    # subtotal header (its scores are the sum of its sub-rows, so we skip them).
    just_saw_challenge_header = False

    for _, row in df.iterrows():
        rule_val = row['Rule']
        rule_str = str(rule_val).strip() if pd.notna(rule_val) else ''
        if rule_str == '':
            just_saw_challenge_header = False
            current_event = None
            continue

        rule_upper = rule_str.upper()

        if 'TOTAL POINTS' in rule_upper:
            for inst in institutions:
                v = row[inst]
                try:
                    totals[inst] = float(v) if pd.notna(v) and str(v).strip() != '' else 0.0
                except (ValueError, TypeError):
                    totals[inst] = 0.0
            continue

        if rule_upper.startswith('RANKING'):
            continue

        if _all_inst_empty(row, institutions):
            if rule_str == rule_str.upper():
                # All-caps = top-level challenge section header (e.g. URBAN CHALLENGE)
                current_challenge = {'name': rule_str, 'events': []}
                challenges.append(current_challenge)
                current_event = None
                just_saw_challenge_header = True
            else:
                # Mixed-case empty row = explicit event name label
                current_event = {'name': rule_str, 'rules': []}
                if current_challenge is not None:
                    current_challenge['events'].append(current_event)
                just_saw_challenge_header = False
        else:
            # Row has numeric data
            scores = {}
            for inst in institutions:
                v = row[inst]
                try:
                    scores[inst] = float(v) if pd.notna(v) and str(v).strip() != '' else 0.0
                except (ValueError, TypeError):
                    scores[inst] = 0.0

            if just_saw_challenge_header:
                # This is an event subtotal row (e.g. "Chancellor Challenge = 76").
                # Its value is the sum of its sub-rows — don't add it to scores.
                current_event = {'name': rule_str, 'rules': []}
                if current_challenge is not None:
                    current_challenge['events'].append(current_event)
                just_saw_challenge_header = False
            else:
                rule_entry = {'label': rule_str, 'scores': scores}
                if current_event is not None:
                    current_event['rules'].append(rule_entry)
                elif current_challenge is not None:
                    direct_event = {'name': '', 'rules': [], '_direct': True}
                    current_challenge['events'].append(direct_event)
                    current_event = direct_event
                    current_event['rules'].append(rule_entry)

                # After 'Event Win Points' the next data row starts the next event
                if rule_str == 'Event Win Points':
                    just_saw_challenge_header = True
                    current_event = None

    # Sum leaf rule scores to get calculated totals
    calculated_totals = {inst: 0.0 for inst in institutions}
    for challenge in challenges:
        for event in challenge['events']:
            for rule in event['rules']:
                for inst in institutions:
                    v = rule['scores'].get(inst, 0.0)
                    if isinstance(v, float) and np.isnan(v):
                        v = 0.0
                    calculated_totals[inst] += v
    calculated_totals = {inst: round(v, 2) for inst, v in calculated_totals.items()}

    # Build rankings
    sorted_insts = sorted(institutions, key=lambda i: calculated_totals[i], reverse=True)
    calculated_rankings = {}
    rank = 1
    for i, inst in enumerate(sorted_insts):
        if i > 0 and calculated_totals[inst] == calculated_totals[sorted_insts[i - 1]]:
            calculated_rankings[inst] = calculated_rankings[sorted_insts[i - 1]]
        else:
            calculated_rankings[inst] = rank
        rank += 1

    return {
        'institutions': institutions,
        'challenges': challenges,
        'totals': totals,
        'calculated_totals': calculated_totals,
        'calculated_rankings': calculated_rankings,
    }


def _doc_season_year(doc):
    """
    Resolve which season year a confirmed ScoreDocument belongs to.
 
    Priority order:
      1. Explicit seasonID FK on the document.
      2. Year written in cell A1 of the spreadsheet (e.g. "Season: 2025").
      3. Year of uploadedOn  ← LAST, because a 2025 spreadsheet may be
         uploaded in 2026.
    """
    # 1. Explicit FK
    if hasattr(doc, 'seasonID') and doc.seasonID:
        season = db.session.get(Season, doc.seasonID)
        if season:
            return season.year
 
    # 2. Read year from the spreadsheet's first cell (e.g. "Season: 2025")
    if doc.fileData:
        try:
            ext = os.path.splitext(doc.originalFilename)[1].lower()
            buf = io.BytesIO(doc.fileData)
            if ext in ('.xlsx', '.xls'):
                from openpyxl import load_workbook
                wb = load_workbook(buf, data_only=True, read_only=True)
                ws = wb.active
                first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                cell_a1 = str(first_row[0]).strip() if first_row and first_row[0] is not None else ''
            else:
                df_raw = pd.read_csv(buf, header=None, nrows=1)
                cell_a1 = str(df_raw.iloc[0, 0]).strip()

            m = re.search(r'\b(20\d{2})\b', cell_a1)
            if m:
                year = int(m.group(1))
                season = Season.query.filter_by(year=year).first()
                return season.year if season else year
        except Exception:
            pass

    # 3. Upload-year fallback
    if doc.uploadedOn:
        upload_year = doc.uploadedOn.year
        season = Season.query.filter_by(year=upload_year).first()
        if season:
            return season.year
        return upload_year

    return None


# ---------------------------------------------------------------------------
# Page route
# ---------------------------------------------------------------------------

@leaderboard_views.route('/leaderboard', methods=['GET'])
def get_leaderboard_page():
    return render_template('user/leaderboard.html')


# ---------------------------------------------------------------------------
# API — leaderboard data from confirmed documents
# ---------------------------------------------------------------------------

@leaderboard_views.route('/api/leaderboard', methods=['GET'])
def get_leaderboard_api():
    """
    Returns aggregated leaderboard data from the most recent season
    that has at least one confirmed ScoreDocument.

    Query params:
        year (int, optional) – override to fetch a specific season year
        event (str, optional) – challenge name filter (case-insensitive)

    Response shape:
    {
        "season_year": 2025,
        "available_years": [2025, 2024],
        "challenges": ["Overall", "100m Sprint", ...],
        "leaderboard": [
            {
                "rank": 1,
                "institution": "UWI St. Augustine",
                "total_points": 1482.0,
                "challenge_points": {
                    "100m Sprint": 320.0,
                    ...
                }
            },
            ...
        ]
    }
    """
    # -- Gather all confirmed documents
    confirmed_docs = db.session.scalars(
        db.select(ScoreDocument).filter_by(confirmed=True)
    ).all()

    if not confirmed_docs:
        return jsonify({
            "season_year": None,
            "available_years": [],
            "challenges": [],
            "leaderboard": [],
            "message": "No confirmed results yet."
        })

    # -- Map each doc to its season year
    docs_by_year: dict[int, list] = {}
    for doc in confirmed_docs:
        year = _doc_season_year(doc)
        if year is None:
            continue
        docs_by_year.setdefault(year, []).append(doc)

    if not docs_by_year:
        return jsonify({
            "season_year": None,
            "available_years": [],
            "challenges": [],
            "leaderboard": [],
            "message": "No confirmed results could be matched to a season."
        })

    available_years = sorted(docs_by_year.keys(), reverse=True)

    # -- Determine target year
    try:
        requested_year = int(request.args.get('year', 0))
    except (ValueError, TypeError):
        requested_year = 0

    if requested_year and requested_year in docs_by_year:
        target_year = requested_year
    else:
        target_year = available_years[0]

    target_docs = docs_by_year[target_year]

    agg: dict[str, dict[str, float]] = {}
    challenge_names: list[str] = []

    for doc in target_docs:
        parsed = _parse_document(doc)
        if parsed is None:
            continue

        for challenge in parsed['challenges']:
            c_name = challenge['name'].strip()
            if c_name and c_name not in challenge_names:
                challenge_names.append(c_name)

            for inst in parsed['institutions']:
                if inst not in agg:
                    agg[inst] = {}

                challenge_pts = 0.0
                for event in challenge['events']:
                    for rule in event['rules']:
                        v = rule['scores'].get(inst, 0.0)
                        if isinstance(v, float) and np.isnan(v):
                            v = 0.0
                        challenge_pts += v

                agg[inst][c_name] = round(
                    agg[inst].get(c_name, 0.0) + challenge_pts, 2
                )

    if not agg:
        return jsonify({
            "season_year": target_year,
            "available_years": available_years,
            "challenges": [],
            "leaderboard": [],
            "message": "Documents found but could not be parsed."
        })

    totals = {inst: round(sum(pts.values()), 2) for inst, pts in agg.items()}
    sorted_insts = sorted(totals.keys(), key=lambda i: totals[i], reverse=True)

    leaderboard = []
    rank = 1
    for i, inst in enumerate(sorted_insts):
        if i > 0 and totals[inst] == totals[sorted_insts[i - 1]]:
            assigned_rank = leaderboard[-1]['rank']
        else:
            assigned_rank = rank
        rank += 1

        inst_name = f"Institution {inst.split('_')[1]}" if inst.startswith('Unnamed') else inst
        leaderboard.append({
            "rank": assigned_rank,
            "institution": inst_name,
            "total_points": totals[inst],
            "challenge_points": agg[inst],
        })

    # -- Optional challenge filter for the event tab
    event_filter = request.args.get('event', '').strip().lower()
    if event_filter and event_filter != 'overall':
        # Re-rank by challenge-specific points
        filtered = []
        for entry in leaderboard:
            matched_key = next(
                (k for k in entry['challenge_points']
                 if k.lower() == event_filter or event_filter in k.lower()),
                None
            )
            pts = entry['challenge_points'].get(matched_key, 0.0) if matched_key else 0.0
            filtered.append({**entry, "filtered_points": pts})

        filtered.sort(key=lambda x: x['filtered_points'], reverse=True)
        rank = 1
        for i, entry in enumerate(filtered):
            if i > 0 and entry['filtered_points'] == filtered[i - 1]['filtered_points']:
                entry['rank'] = filtered[i - 1]['rank']
            else:
                entry['rank'] = rank
            rank += 1

        leaderboard = filtered

    return jsonify({
        "season_year": target_year,
        "available_years": available_years,
        "challenges": challenge_names,
        "leaderboard": leaderboard,
    })