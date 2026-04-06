import os
from datetime import date, datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, abort
from flask_jwt_extended import jwt_required, current_user

from flask import send_file
import openpyxl
from io import BytesIO

from App.models import Event, Institution, Season
from App.models.challenge import Challenge
from App.database import db

from App.controllers import (
    upload_score_document,
    get_my_score_documents,
    delete_score_document
)

scoretaker_views = Blueprint('scoretaker_views', __name__, url_prefix='/scoretaker')

ALLOWED_EXTENSIONS = {
    ".xml", ".xlsx", ".xls", ".xlsm", ".xlsb",
    ".xlam", ".xltx", ".xltm", ".csv", ".ods"
}


def _allowed(filename: str) -> bool:
    return os.path.splitext(filename.lower())[1] in ALLOWED_EXTENSIONS


# ── Dashboard ─────────────────────────────────────────────────────────────────

@scoretaker_views.route('/', methods=['GET'])
@jwt_required()
def scoretaker_dashboard():
    documents = get_my_score_documents(current_user.userID)
    today = date.today()

    total_uploads  = len(documents)
    uploaded_today = len([
        d for d in documents
        if d.uploadedOn and d.uploadedOn.date() == today
    ])

    seasons = Season.query.order_by(Season.year.desc()).all()
    current_year   = datetime.now().year
    current_season = Season.query.filter_by(year=current_year).first()

    return render_template(
        'scoretaker/scoretaker.html',
        total_uploads=total_uploads,
        uploaded_today=uploaded_today,
        user=current_user,
        seasons=seasons,
        current_season_id=current_season.seasonID if current_season else (seasons[0].seasonID if seasons else None),
    )


# ── Upload Scores ─────────────────────────────────────────────────────────────

@scoretaker_views.route('/upload', methods=['GET', 'POST'])
@jwt_required()
def upload_scores():
    if request.method == 'POST':
        files = request.files.getlist('files')

        if not files or files[0].filename == '':
            flash("No files selected.", "error")
            return redirect(url_for('scoretaker_views.upload_scores'))

        upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
        os.makedirs(upload_folder, exist_ok=True)

        success_count = 0
        error_count   = 0

        for file in files:
            try:
                if not _allowed(file.filename):
                    flash(f"'{file.filename}' is not a supported format.", "error")
                    error_count += 1
                    continue

                upload_score_document(
                    userID=current_user.userID,
                    file_storage=file,
                    upload_folder=upload_folder
                )
                success_count += 1

            except Exception as e:
                flash(f"Error uploading '{file.filename}': {e}", "error")
                error_count += 1

        if success_count:
            flash(f"{success_count} file(s) uploaded successfully.", "success")
        if error_count:
            flash(f"{error_count} file(s) failed to upload.", "error")

        return redirect(url_for('scoretaker_views.archives'))

    return render_template('scoretaker/uploadScores.html', user=current_user)


# ── Archives ──────────────────────────────────────────────────────────────────

@scoretaker_views.route('/archives', methods=['GET'])
@jwt_required()
def archives():
    documents = get_my_score_documents(current_user.userID)

    docs_data = []
    for d in documents:
        ext = os.path.splitext(d.originalFilename)[1].lstrip('.').upper() if d.originalFilename else ''

        uploaded_by = "—"
        if d.scoretaker and d.scoretaker.user and d.scoretaker.user.username:
            uploaded_by = d.scoretaker.user.username

        docs_data.append({
            "id":             d.documentID,
            "filename":       d.originalFilename or '—',
            "storedFilename": d.storedFilename,
            "uploadedAt":     d.uploadedOn.strftime("%b %d, %Y · %H:%M") if d.uploadedOn else "—",
            "uploadedAtRaw":  d.uploadedOn.isoformat() if d.uploadedOn else "",
            "fileType":       ext or "FILE",
            "uploadedBy":     uploaded_by,
            "viewUrl":        url_for('scoretaker_views.view_document', documentID=d.documentID),
            "deleteUrl":      url_for('scoretaker_views.delete_document', documentID=d.documentID),
        })

    docs_data.sort(key=lambda x: x["uploadedAtRaw"], reverse=True)

    return render_template(
        'scoretaker/archives.html',
        documents=docs_data,
        total=len(docs_data),
        user=current_user
    )


# ── View / Download ───────────────────────────────────────────────────────────

@scoretaker_views.route('/document/<int:documentID>', methods=['GET'])
@jwt_required()
def view_document(documentID):
    from App.models import ScoreDocument
    doc = ScoreDocument.query.filter_by(
        documentID=documentID,
        scoretakerID=current_user.userID
    ).first_or_404()

    if not os.path.exists(doc.storedPath):
        abort(404)

    return send_file(
        doc.storedPath,
        as_attachment=False,
        download_name=doc.originalFilename
    )


# ── Delete ────────────────────────────────────────────────────────────────────

@scoretaker_views.route('/delete/<int:documentID>', methods=['POST'])
@jwt_required()
def delete_document(documentID):
    try:
        delete_score_document(current_user.userID, documentID)
        flash("Document deleted.", "success")
    except Exception as e:
        flash(str(e), "error")

    return redirect(url_for('scoretaker_views.archives'))


# ── Debug: inspect ruleType values (remove after confirming) ─────────────────

@scoretaker_views.route("/debug-rules")
@jwt_required()
def debug_rules():
    from App.models import PointsRules
    rows = PointsRules.query.all()
    lines = [
        f"pointsID={r.pointsID} | eventID={r.eventID} | ruleType={r.ruleType!r} | category={r.category!r} | label={r.label!r} | points={r.points}"
        for r in rows
    ]
    return "<pre>" + "\n".join(lines) + "</pre>"


# ── Score Template Download ───────────────────────────────────────────────────

@scoretaker_views.route("/template")
@jwt_required()
def download_template():
    """
    Generate a season-filtered Excel score template.

    Sheet structure per event:
      [Event Name row]              ← auto-SUM of all sub-rows below it
        [Category Name row]         ← input cells (one per team category)
        ...
        [Event Win Points row]      ← input cells (only if win points configured)
      [gap]
    [TOTAL POINTS row]  (SUM of every sub-row across all events)
    [RANKING row]       (RANK formula)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Resolve season ────────────────────────────────────────────────────────
    season_id = request.args.get("season_id", type=int)
    if season_id:
        selected_season = db.session.get(Season, season_id)
        if not selected_season:
            flash("Season not found.", "error")
            selected_season = None
    else:
        selected_season = Season.query.filter_by(year=datetime.now().year).first()

    if selected_season:
        events       = Event.query.filter_by(seasonID=selected_season.seasonID).order_by(Event.eventName.asc()).all()
        challenges   = Challenge.query.filter_by(seasonID=selected_season.seasonID).order_by(Challenge.challengeName.asc()).all()
        season_label = str(selected_season.year)
    else:
        events       = Event.query.order_by(Event.eventName.asc()).all()
        challenges   = Challenge.query.order_by(Challenge.challengeName.asc()).all()
        season_label = "All"

    institutions = Institution.query.order_by(Institution.insName.asc()).all()
    n_inst       = len(institutions)
    total_cols   = n_inst + 1   # col A = label; cols B..N = institutions

    # ── Colour palette ────────────────────────────────────────────────────────
    C_SEASON_BG  = "050E2B"
    C_INST_HDR   = "0A2A66"
    C_CHALLENGE  = "0B3D91"
    C_EVENT      = "1565C0"   # event header row (bold, shows SUM)
    C_CAT_BG     = "1976D2"   # team category sub-rows (slightly lighter blue)
    C_CAT_INPUT  = "E3F2FD"   # input cells for category rows
    C_WIN_PTS    = "E8F5E9"   # event win points sub-row label
    C_WIN_INPUT  = "F1F8E9"   # input cells for win points row
    C_TOTAL_BG   = "1C1C1C"
    C_RANK_BG    = "424242"
    C_WHITE      = "FFFFFF"
    C_BORDER     = "90CAF9"
    C_CAT_BORDER = "64B5F6"
    C_WIN_BORDER = "A5D6A7"
    C_WIN_TEXT   = "1B5E20"
    C_CAT_TEXT   = "E3F2FD"

    FONT_NAME = "Arial"

    def fill(h):
        return PatternFill("solid", fgColor=h, start_color=h)

    def fnt(bold=False, italic=False, color=C_WHITE, size=10):
        return Font(name=FONT_NAME, bold=bold, italic=italic, color=color, size=size)

    def bdr(color=C_BORDER):
        s = Side(border_style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def align_center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def align_left(indent=1):
        return Alignment(horizontal="left", vertical="center", indent=indent)

    # ── Workbook / column widths ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores Template"

    ws.column_dimensions["A"].width = 36
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    last_inst_col = get_column_letter(total_cols)
    row = 1

    # ── Helpers to extract rules from an event via PointsRules ───────────────

    # ruleType values written by admin_event.js:
    #   placement-based win rules -> 'individual'  (placement + label + points)
    #   team category rules       -> 'team'         (category + label + points)
    # Update these constants if your JS writes different strings.
    RTYPE_TEAM = 'team'
    RTYPE_WIN  = 'individual'

    def _get_team_categories(event):
        """
        Return a deduplicated ordered list of team category name strings
        from the event's PointsRules rows where ruleType == RTYPE_TEAM.
        """
        seen = []
        for rule in (event.points_rules or []):
            if rule.ruleType == RTYPE_TEAM and rule.category and rule.category not in seen:
                seen.append(rule.category)
        return seen

    def _has_event_win_points(event):
        """
        Return True if the event has any PointsRules rows
        where ruleType == RTYPE_WIN.
        """
        return any(r.ruleType == RTYPE_WIN for r in (event.points_rules or []))

    # ── Row writers ───────────────────────────────────────────────────────────

    def write_season_banner():
        nonlocal row
        ws.row_dimensions[row].height = 22
        c = ws.cell(row=row, column=1, value=f"Season: {season_label}")
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color="C9A6FF")
        c.fill      = fill(C_SEASON_BG)
        c.alignment = align_left()
        for col in range(2, total_cols + 1):
            ws.cell(row=row, column=col).fill = fill(C_SEASON_BG)
        row += 1

    def write_inst_header():
        nonlocal row
        ws.row_dimensions[row].height = 30
        lbl = ws.cell(row=row, column=1, value="Event / Institution")
        lbl.font      = fnt(bold=True, color="C9A6FF")
        lbl.fill      = fill(C_INST_HDR)
        lbl.alignment = align_left()
        for i, inst in enumerate(institutions, start=2):
            c = ws.cell(row=row, column=i, value=inst.insName)
            c.font      = fnt(bold=True)
            c.fill      = fill(C_INST_HDR)
            c.alignment = align_center(wrap=True)
        row += 1

    def write_challenge_banner(name):
        nonlocal row
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=1, value=name.upper())
        c.font      = fnt(bold=True, size=12)
        c.fill      = fill(C_CHALLENGE)
        c.alignment = align_left()
        for col in range(2, total_cols + 1):
            ws.cell(row=row, column=col).fill = fill(C_CHALLENGE)
        row += 1

    def write_event_rows(event):
        """
        Write one event header row + one sub-row per team category +
        optionally one Event Win Points sub-row.

        The event header row shows a live SUM of all its sub-rows.
        Returns a list of sub-row numbers (these feed the grand TOTAL formula).
        """
        nonlocal row

        team_categories = _get_team_categories(event)
        has_win_pts     = _has_event_win_points(event)

        # ── Reserve the header row; we'll write its SUM formula after
        #    the sub-rows so we know their row numbers. ────────────────────────
        header_row_num = row
        ws.row_dimensions[header_row_num].height = 22
        row += 1

        sub_rows = []   # row numbers of every input sub-row for this event

        # ── Team category sub-rows ────────────────────────────────────────────
        for cat_name in team_categories:
            ws.row_dimensions[row].height = 17
            lbl = ws.cell(row=row, column=1, value=cat_name)
            lbl.font      = Font(name=FONT_NAME, size=9, italic=True, color=C_CAT_TEXT)
            lbl.fill      = fill(C_CAT_BG)
            lbl.alignment = align_left(indent=4)
            for col in range(2, total_cols + 1):
                c = ws.cell(row=row, column=col)
                c.fill          = fill(C_CAT_INPUT)
                c.border        = bdr(C_CAT_BORDER)
                c.alignment     = align_center()
                c.number_format = "0.00"
            sub_rows.append(row)
            row += 1

        # ── Event Win Points sub-row (only if configured) ─────────────────────
        if has_win_pts:
            ws.row_dimensions[row].height = 17
            lbl = ws.cell(row=row, column=1, value="Event Win Points")
            lbl.font      = Font(name=FONT_NAME, size=9, italic=True, color=C_WIN_TEXT)
            lbl.fill      = fill(C_WIN_PTS)
            lbl.alignment = align_left(indent=4)
            for col in range(2, total_cols + 1):
                c = ws.cell(row=row, column=col)
                c.fill          = fill(C_WIN_INPUT)
                c.border        = bdr(C_WIN_BORDER)
                c.alignment     = align_center()
                c.number_format = "0.00"
            sub_rows.append(row)
            row += 1

        # ── Now write the event header row with its SUM formula ───────────────
        lbl = ws.cell(row=header_row_num, column=1, value=event.eventName)
        lbl.font      = fnt(bold=True, size=10)
        lbl.fill      = fill(C_EVENT)
        lbl.alignment = align_left(indent=2)

        for col in range(2, total_cols + 1):
            col_letter = get_column_letter(col)
            c = ws.cell(row=header_row_num, column=col)
            c.fill          = fill(C_EVENT)
            c.font          = fnt(bold=True, size=10)
            c.alignment     = align_center()
            c.number_format = "0.00"
            if sub_rows:
                # SUM of all sub-rows for this column
                sum_parts = "+".join(f"{col_letter}{r}" for r in sub_rows)
                c.value = f"={sum_parts}"
            # If an event somehow has no sub-rows, leave the header blank

        return sub_rows   # caller adds these to the grand total

    def write_gap():
        nonlocal row
        ws.row_dimensions[row].height = 6
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).fill = fill("E8F4FD")
        row += 1

    # ── Build sheet ───────────────────────────────────────────────────────────
    # score_rows collects every input sub-row across all events.
    # The grand TOTAL formula sums these — event header rows are excluded
    # because they are derived (SUM of their own sub-rows), so including
    # them would double-count.
    score_rows          = []
    challenge_event_ids = set()

    write_season_banner()

    for ch in challenges:
        ch_events = [
            ev for ev in ch.events
            if not selected_season or ev.seasonID == selected_season.seasonID
        ]
        if not ch_events:
            continue

        write_inst_header()
        write_challenge_banner(ch.challengeName)

        for ev in ch_events:
            challenge_event_ids.add(ev.eventID)
            sub_rows = write_event_rows(ev)
            score_rows.extend(sub_rows)

        write_gap()

    # Orphan events (not attached to any challenge in this season)
    orphans = [e for e in events if e.eventID not in challenge_event_ids]
    if orphans:
        write_inst_header()
        write_challenge_banner("Other Events")
        for ev in orphans:
            sub_rows = write_event_rows(ev)
            score_rows.extend(sub_rows)
        write_gap()

    write_gap()

    # ── TOTAL POINTS + RANKING ────────────────────────────────────────────────
    total_row_num = row
    rank_row_num  = row + 1

    ws.row_dimensions[total_row_num].height = 24
    ws.row_dimensions[rank_row_num].height  = 24

    rank_range = f"$B${total_row_num}:${last_inst_col}${total_row_num}"

    for r_num, label, bg in [
        (total_row_num, "TOTAL POINTS", C_TOTAL_BG),
        (rank_row_num,  "RANKING",      C_RANK_BG),
    ]:
        lbl = ws.cell(row=r_num, column=1, value=label)
        lbl.font      = fnt(bold=True, size=11)
        lbl.fill      = fill(bg)
        lbl.alignment = align_left()
        for col in range(2, total_cols + 1):
            c           = ws.cell(row=r_num, column=col)
            c.fill      = fill(bg)
            c.font      = fnt(bold=True, size=11)
            c.alignment = align_center()

    if score_rows:
        for col_idx in range(2, total_cols + 1):
            col_letter = get_column_letter(col_idx)
            sum_parts  = "+".join(f"{col_letter}{r}" for r in score_rows)
            ws.cell(row=total_row_num, column=col_idx, value=f"={sum_parts}")
            ws.cell(row=rank_row_num,  column=col_idx,
                    value=f"=RANK({col_letter}{total_row_num},{rank_range},0)")

    # ── Stream ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"score_template_{season_label}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )